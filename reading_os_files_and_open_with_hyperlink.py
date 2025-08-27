import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- Config ---
# Change this to your target folder
folder_path = r"your folder path"

# Output Excel file path (change this to any directory you want)
output_dir = r"your folder path"   # Replace with your desired folder
os.makedirs(output_dir, exist_ok=True)  # Create folder if not exists
output_excel = os.path.join(output_dir, "file_list.xlsx")

# --- Collect file paths ---
file_list = []
for root, dirs, files in os.walk(folder_path):
    for file in files:
        file_path = os.path.join(root, file)
        file_list.append({
            "File Name": file,
            "Folder Path": root,
            "Hyperlink": file_path  # store actual path
        })

# --- Create DataFrame ---
df = pd.DataFrame(file_list)

# --- Export to Excel ---
df.to_excel(output_excel, index=False)

# --- Apply hyperlink formula & style ---
wb = load_workbook(output_excel)
ws = wb.active

# Find "Hyperlink" column index
hyperlink_col = None
file_name_col = None
for idx, col in enumerate(ws[1], start=1):
    if col.value == "Hyperlink":
        hyperlink_col = idx
    elif col.value == "File Name":
        file_name_col = idx

# Apply hyperlink formula and style
if hyperlink_col and file_name_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=hyperlink_col)
        file_path = cell.value
        file_name = ws.cell(row=row, column=file_name_col).value
        if file_path and file_name:  
            cell.value = f'=HYPERLINK("{file_path}", "{file_name}")'
            cell.font = Font(color="0000FF", underline="single")  # Blue & underlined

wb.save(output_excel)

print(f"✅ Exported {len(df)} files to {output_excel}")
